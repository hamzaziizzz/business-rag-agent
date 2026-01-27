from __future__ import annotations

from io import BytesIO
from typing import Any

from src.rag.types import Document


class XlsxLoaderError(RuntimeError):
    pass


def load_xlsx_bytes(data: bytes, doc_id: str, source: str) -> Document:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise XlsxLoaderError("openpyxl is required to load XLSX files") from exc

    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    parts: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        parts.append(f"Sheet: {sheet_name}")
        for row in sheet.iter_rows(values_only=True):
            row_values = [_format_cell(value) for value in row]
            if any(row_values):
                parts.append(", ".join(row_values))
    content = "\n".join(parts).strip()
    return Document(doc_id=doc_id, content=content, metadata={"source": source})


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
